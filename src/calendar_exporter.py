"""
Calendar Exporter Module
Export name days and calendars to Excel format
"""

from datetime import datetime, timedelta
from pathlib import Path
import json
try:
    from .path_manager import get_paths
except ImportError:
    from path_manager import get_paths

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class CalendarExporter:
    """Export Kazamias calendar data to Excel"""

    def __init__(self, db, output_dir: str = None):
        """Initialize exporter"""
        self.db = db

        if output_dir is None:
            output_dir = get_paths().app_output_dir
        else:
            output_dir = Path(output_dir)

        output_dir.mkdir(parents=True, exist_ok=True)
        self.output_dir = output_dir

    def export_daily_message_txt(self, message_text: str, date_str: str = None) -> bool:
        """Export daily message to TXT file"""

        try:
            if date_str is None:
                date_str = datetime.now().strftime('%Y-%m-%d')

            output_file = self.output_dir / f"KazaALKIS_Message_{date_str}.txt"

            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(f"KazaALKIS Daily Message - {date_str}\n")
                f.write("="*70 + "\n\n")
                f.write(message_text)
                f.write(f"\n\nGenerated: {datetime.now().isoformat()}\n")

            print(f"✓ Message exported to {output_file}")
            return True

        except Exception as e:
            print(f"✗ Error exporting message: {e}")
            return False

    def export_contacts_xlsx(self) -> bool:
        """Export contacts to Excel"""

        if not HAS_OPENPYXL:
            print("✗ openpyxl not installed")
            return False

    def export_logs_xlsx(self) -> bool:
        """Export message audit logs with masked recipients."""
        if not HAS_OPENPYXL:
            print("openpyxl not installed")
            return False
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Message Logs"
            ws.append(["Timestamp", "Recipient", "Status", "Delivery Method", "Error"])
            self.db.cursor.execute("""
                SELECT m.sent_at, c.phone_masked, m.status, m.delivery_method, m.error_message
                FROM message_logs m
                JOIN contacts c ON c.id = m.contact_id
                ORDER BY m.sent_at DESC
            """)
            for row in self.db.cursor.fetchall():
                ws.append(list(row))
            output_file = self.output_dir / f"KazaALKIS_Logs_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(output_file)
            print(f"Message logs exported to {output_file}")
            return True
        except Exception as e:
            print(f"Error exporting logs: {e}")
            return False

    def export_monthly_namedays_xlsx(self, year: int, month: int) -> bool:
        """Export one month's nameday calendar."""
        if not HAS_OPENPYXL:
            print("openpyxl not installed")
            return False
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Name Days"
            ws.append(["Date", "Names", "Saint"])
            prefix = f"{year:04d}-{month:02d}"
            self.db.cursor.execute("""
                SELECT date, names, saint FROM name_days
                WHERE date LIKE ? ORDER BY date
            """, (prefix + "%",))
            for row in self.db.cursor.fetchall():
                ws.append(list(row))
            output_file = self.output_dir / f"KazaALKIS_NameDays_{prefix}.xlsx"
            wb.save(output_file)
            print(f"Nameday calendar exported to {output_file}")
            return True
        except Exception as e:
            print(f"Error exporting namedays: {e}")
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contacts"

            headers = ['ID', 'Name', 'Phone (Masked)', 'Language', 'Timezone', 'Active']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col).value = header

            self.db.cursor.execute("""
                SELECT id, name, phone_masked, language, timezone, is_active
                FROM contacts
                ORDER BY name
            """)

            contacts = self.db.cursor.fetchall()

            for row_idx, contact in enumerate(contacts, 2):
                for col_idx, value in enumerate(contact, 1):
                    ws.cell(row=row_idx, column=col_idx).value = value

            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 8

            output_file = self.output_dir / f"KazaALKIS_Contacts_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(output_file)

            print(f"✓ Contacts exported to {output_file}")
            return True

        except Exception as e:
            print(f"✗ Error exporting contacts: {e}")
            return False

if __name__ == "__main__":
    print("Calendar exporter module loaded")
